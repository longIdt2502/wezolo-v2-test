import random
from datetime import datetime

import openpyxl
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.http import HttpResponse

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny

from django.db.models import OuterRef, F, Q

from common.core.subquery import *
from common.s3 import AwsS3
from utils.convert_response import convert_response
from zalo.utils import convert_phone

from .models import Customer, CustomerUserZalo, CustomerImport
from user.models import User
from workspace.models import Workspace
from tags.models import Tag, TagCustomer
from zalo.models import UserZalo, ZaloOA
from progress.models import ProgressTagCustomer


class CustomerCreate(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        with transaction.atomic():
            try:
                user = request.user
                data = request.data.copy()
                ws = data.get('workspace')
                if not ws:
                    raise Exception('Yêu cầu thông tin workspace')
                oas = data.get('oa', [])
                phone = data.get('phone')
                phone = convert_phone(phone)
                customer = Customer.objects.filter(workspace_id=ws, phone=phone).first()
                if customer:
                    for oa in oas:
                        customer_user_zalo = CustomerUserZalo.objects.filter(
                            customer=customer, oa_id=oa
                        )
                        if customer_user_zalo:
                            raise Exception('Khách hàng đã tồn tại trong Workspace và Zalo Oa')
                        else:
                            CustomerUserZalo.objects.create(
                                customer=customer,
                                oa_id=oa,
                            )
                            return convert_response('success', 200, data=customer.id)

                customer = Customer.objects.create(
                    phone=phone,
                    prefix_name=data.get('prefix_name'),
                    email=data.get('email'),
                    address=data.get('address'),
                    gender=data.get('gender'),
                    note=data.get('note'),
                    workspace_id=ws,
                    created_by=user,
                )
                for oa in oas:
                    user_zalo = UserZalo.objects.filter(phone=phone, oa_id=oa).first()
                    if user_zalo:
                        CustomerUserZalo.objects.create(
                            customer=customer,
                            oa_id=oa,
                            user_zalo=user
                        )
                return convert_response('success', 200, data=customer.id)

            except Exception as e:
                return convert_response(str(e), 400)


class CustomerList(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        data = request.GET.copy()
        page_size = int(data.get('page_size', 20))
        offset = (int(data.get('page', 1)) - 1) * page_size
        ws = data.get('workspace')
        if not ws:
            raise Exception('Yêu cầu thông tin Workspace')
        customer = Customer.objects.filter(workspace_id=ws).order_by('-id')

        oa_id = data.get('oa')
        if oa_id:
            customer_user_zalo = CustomerUserZalo.objects.filter(oa_id=oa_id).values_list('customer_id', flat=True)
            customer = customer.filter(id__in=customer_user_zalo)

        gender = data.get('gender')
        if gender:
            customer = customer.filter(customer__gender=gender)

        search = data.get('search')
        if search:
            customer = customer.filter(Q(prefix_name__icontains=search) | Q(phone__icontains=search))

        oa_subquery = SubqueryJsonAgg(
            CustomerUserZalo.objects.filter(user_zalo__is_follower=True, customer_id=OuterRef('id')).exclude(oa_id=None).values().annotate(
                oa_name=F('oa__oa_name'),
                oa_avatar=F('oa__oa_avatar')
            )
        )

        oa_unfollow_subquery = SubqueryJsonAgg(
            CustomerUserZalo.objects.filter(customer_id=OuterRef('id')).filter(Q(user_zalo__is_follower=False) | Q(user_zalo=None)).exclude(oa_id=None).values().annotate(
                oa_name=F('oa__oa_name'),
                oa_avatar=F('oa__oa_avatar')
            )
        )

        tag_subquery = SubqueryJsonAgg(
            TagCustomer.objects.filter(customer_id=OuterRef('id')).values(
                'tag_id', 'tag__title', 'tag__color_border', 'tag__color_border', 'tag__color_fill'
            )
        )

        tag_id = data.get('tags')
        if tag_id:
            tag_id = tag_id.split(',')
            tag_cus = TagCustomer.objects.filter(tag_id__in=tag_id)
            cus = tag_cus.values_list('customer_id', flat=True)
            customer = customer.filter(id__in=cus)

        progress_tag_subquery = SubqueryJson(
            ProgressTagCustomer.objects.filter(customer_id=OuterRef('id')).values(
                'tag__progress_id' , 'tag_id',
                'tag__title', 'tag__color_text', 'tag__color_fill', 'tag__color_border',
                'tag__type', 'tag__progress__title'
            )[:1]
        )

        user_zalo_query = SubqueryJson(
            CustomerUserZalo.objects.filter(customer_id=OuterRef('id')).values(
                'user_zalo__id', 'user_zalo__name', 'user_zalo__phone', 'user_zalo__avatar_small', 
                'user_zalo__avatar_big', 'user_zalo__user_zalo_id'
            )[:1]
        )

        total = customer.count()
        customer = customer[offset: offset + page_size].values().annotate(
            oa_follow=oa_subquery,
            oa_unfollow=oa_unfollow_subquery,
            created_user=SubqueryJson(
                User.objects.filter(id=OuterRef('created_by_id')).values(
                    'id', 'phone', 'full_name', 'avatar'
                )[:1]
            ),
            tag=tag_subquery,
            progress=progress_tag_subquery,
            user_zalo=user_zalo_query,
        )

        return convert_response('success', 200, data=customer, total=total)


class CustomerDetail(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            tags_subquery = SubqueryJsonAgg(
                Tag.objects.filter(oa_id=OuterRef('oa_id')).values(
                    'id', 'title', 'color_text', 'color_fill', 'color_border'
                )
            )
            customer = Customer.objects.get(id=pk)
            customer = customer.to_json()

            progress_tag = ProgressTagCustomer.objects.filter(customer_id=customer['id']).values(
                'tag__progress__title', 'tag__type', 'tag__title', 'tag__color_text', 'tag__color_fill', 'tag__color_border'
            ).first()

            customer['progress'] = progress_tag

            oa_follow = CustomerUserZalo.objects.filter(customer_id=pk)
            customer['oa_follow'] = oa_follow.values(
                'oa__oa_name', 'oa__oa_avatar'
            ).annotate(
                tags=tags_subquery
            )

            customer['oa_unfollow'] = ZaloOA.objects.filter(company_id=customer['workspace']['id']).exclude(
                id__in=oa_follow.values_list('oa_id', flat=True)
            ).values('oa_name', 'oa_avatar').annotate(
                tags=SubqueryJsonAgg(
                    Tag.objects.filter(oa_id=OuterRef('id')).values(
                        'id', 'title', 'color_text', 'color_fill', 'color_border'
                    )
                )
            )
            return convert_response('success', 200, data=customer)

        except Exception as e:
            return convert_response(str(e), 400)

    def put(self, request, pk):
        user = request.user
        data = request.data.copy()
        try:
            customer = Customer.objects.get(id=pk)
            customer.prefix_name = data.get('prefix_name', customer.prefix_name)
            customer.phone = data.get('phone', customer.phone)
            customer.email = data.get('email', customer.email)
            customer.address = data.get('address', customer.address)
            customer.gender = data.get('gender', customer.gender)
            customer.birthday = datetime.strptime(data.get('birthday'), "%d/%m/%Y") if data.get('birthday') else None

            oas_add = data.get('oa_add', [])
            for oa in oas_add:
                cuz = CustomerUserZalo.objects.filter(customer=customer, oa_id=oa).first()
                if not cuz:
                    CustomerUserZalo.objects.create(
                        customer=customer,
                        oa_id=oa,
                    )

            oas_remove = data.get('oa_remove', [])
            for oa in oas_remove:
                cuz = CustomerUserZalo.objects.filter(customer=customer, oa_id=oa).first()
                if cuz:
                    cuz.delete()
            
            tags = data.get('tags', [])
            for tag in tags:
                tag_cus = TagCustomer.objects.filter(customer=customer, tag_id=tag).first()
                if not tag_cus:
                    TagCustomer.objects.create(
                        customer=customer,
                        tag_id=tag,
                        created_by=user
                    )
            
            tags = data.get('tags_remove', [])
            for tag in tags:
                tag_cus = TagCustomer.objects.filter(customer=customer, tag_id=tag).first()
                if tag_cus:
                    tag_cus.delete()

            tag_progress = data.get('tag_progress')
            if tag_progress:
                tag_cus = ProgressTagCustomer.objects.filter(tag_id=tag_progress, customer=customer).first()
                if not tag_cus:
                    ProgressTagCustomer.objects.create(
                        tag_id=tag_progress,
                        customer=customer,
                        created_by=user,
                    )
            tag_progress = data.get('tag_progress_remove')
            if tag_progress:
                tag_cus = ProgressTagCustomer.objects.filter(tag_id=tag_progress, customer=customer).first()
                if not tag_cus:
                    tag_cus.delete()

            customer.save()

            return convert_response('success', 200)

        except Exception as e:
            return convert_response(str(e), 400)


class ExportFileImportCustomer(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            data = request.GET.copy()
            workspace = data.get('workspace')
            if not workspace:
                return convert_response('Cần thông tin workspace', 400)
            workspace = Workspace.objects.get(id=workspace)
            oa = ZaloOA.objects.filter(company=workspace)

            list_header = ["STT", "Họ tên khách hàng *\n(2-30)", "Số điện thoại *\n(10)",
                           "Ngày sinh\n(DD/MM/YYYY)", "Địa chỉ\n(500)", "Email (100)",
                           "Giới tính\n(1: nam; 2: nữ)"]
            list_tag_id = [None, None, None, None, None, None, None]
            list_tag = [None, None, None, None, None, None, None]
            list_count_tag_by_oa = []
            for item in oa:
                tags = Tag.objects.filter(oa=item)
                list_count_tag_by_oa.append(tags.count())
                for tag in tags:
                    list_header.append(item.oa_name)
                    list_tag_id.append(tag.id)
                    list_tag.append(tag.title)

            # Tạo workbook và worksheet
            wb = openpyxl.Workbook()

            # Sheet "Danh sách"
            ws1 = wb.active
            ws1.title = "Danh sách"

            # Dữ liệu mẫu từ tệp ban đầu
            data_danhsach = [
                [f"[{workspace.name}]"],
                [None, None, None, "DANH SÁCH KHÁCH HÀNG"],
                [],
                list_header,
                list_tag_id,
                list_tag,
            ]

            # Thêm dữ liệu vào "Danh sách"
            for row in data_danhsach:
                ws1.append(row)

            # Gộp các cột tag có cùng OA vào làm 1 ô đối với hàng chứa OA Name
            index = 8
            for count in list_count_tag_by_oa:
                ws1.merge_cells(start_row=4, end_row=4, start_column=index, end_column=index + count - 1)
                index += count

            # Gộp ô (Merge Cells)
            ws1.merge_cells('A1:B2')
            ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['A1'].font = Font(name='Times New Roman', bold=True, size=11)
            ws1['A1'].fill = PatternFill("solid", fgColor="FFCC00")  # Màu vàng
            ws1.column_dimensions['B'].width = 25
            ws1.column_dimensions['C'].width = 20
            ws1.column_dimensions['D'].width = 20
            ws1.column_dimensions['E'].width = 20
            ws1.column_dimensions['F'].width = 20
            ws1.column_dimensions['G'].width = 15
            ws1.row_dimensions[2].height = 30
            ws1.row_dimensions[4].height = 20
            ws1.row_dimensions[5].height = 20
            ws1.row_dimensions[6].height = 20

            ws1.merge_cells('D2:G2')
            ws1['D2'].alignment = Alignment(horizontal="center", vertical="center")
            ws1['D2'].font = Font(name='Times New Roman', bold=True, size=11)

            ws1.merge_cells('A4:A6')
            ws1.merge_cells('B4:B6')
            ws1.merge_cells('C4:C6')
            ws1.merge_cells('D4:D6')
            ws1.merge_cells('E4:E6')
            ws1.merge_cells('F4:F6')
            ws1.merge_cells('G4:G6')

            header_fill = PatternFill('solid', fgColor="FFF2CC")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Định dạng ô trong sheet "Danh sách"
            for row in ws1.iter_rows(min_row=4, max_row=6):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name='Times New Roman', bold=False, size=11, color="000000")
                    cell.fill = header_fill
                    cell.border = thin_border
            # Định dạng ô trong sheet "Danh sách"
            for row in ws1.iter_rows(min_row=4, max_row=10):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name='Times New Roman', bold=False, size=11, color="000000")
                    cell.border = thin_border

            # Sheet "Hướng dẫn"
            ws2 = wb.create_sheet(title="Hướng dẫn")

            # Dữ liệu mẫu từ tệp ban đầu
            data_danhsach = [
                [f"{workspace.name}"],
                [None, None, None, "DANH SÁCH KHÁCH HÀNG"],
                [],
                list_header,
                list_tag_id,
                list_tag,
                [],
                [],
                [],
                [],
                [],
                [None,
                 "Bắt buộc nhập\nTối thiểu 2 ký tự và tối đa 30 ký tự",
                 "Bắt buộc nhập\nCần thiết lập định dạng văn bản để hiển thị được số 0\nSố điện thoại định dạng 10 số",
                 "Không bắt buộc nhập\nCần nhập đúng định dạng quy định: DD/MM/YYYY = Ngày/Tháng/Năm",
                 "Không bắt buộc nhập\nTối đa 500 ký tự",
                 "Không bắt buộc nhập\nTối đa 100 ký tự",
                 "Không bắt buộc nhập\nChỉ nhập mã tương ứng của giới tính là 1 (Nam) hoặc 2 (Nữ). Không có thông tin, không cần nhập trường này",
                 "Không bắt buộc nhập\nNhập '1' để đánh tag cho khách hàng tương ứng.\nNếu đánh tag của OA nào thì sẽ tự động đánh dấu khách thuộc OA tương ứng",
                 ],
            ]

            # Thêm dữ liệu vào "Danh sách"
            for row in data_danhsach:
                ws2.append(row)

            # Gộp các cột tag có cùng OA vào làm 1 ô đối với hàng chứa OA Name
            index = 8
            for count in list_count_tag_by_oa:
                ws2.merge_cells(start_row=4, end_row=4, start_column=index, end_column=index + count - 1)
                index += count
            ws2.merge_cells(start_row=12, end_row=12, start_column=8, end_column=index - 1)

            # Gộp ô (Merge Cells)
            ws2.merge_cells('A1:B2')
            ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws2['A1'].font = Font(name='Times New Roman', bold=True, size=11)
            ws2['A1'].fill = PatternFill("solid", fgColor="FFCC00")  # Màu vàng
            ws2.column_dimensions['B'].width = 25
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['E'].width = 15
            ws2.column_dimensions['F'].width = 15
            ws2.column_dimensions['G'].width = 15
            ws2.row_dimensions[2].height = 30
            ws2.row_dimensions[4].height = 20
            ws2.row_dimensions[5].height = 20
            ws2.row_dimensions[6].height = 20
            ws2.row_dimensions[12].height = 150

            ws2.merge_cells('D2:G2')
            ws2['D2'].alignment = Alignment(horizontal="center", vertical="center")
            ws2['D2'].font = Font(name='Times New Roman', bold=True, size=11)

            ws2.merge_cells('A4:A6')
            ws2.merge_cells('B4:B6')
            ws2.merge_cells('C4:C6')
            ws2.merge_cells('D4:D6')
            ws2.merge_cells('E4:E6')
            ws2.merge_cells('F4:F6')
            ws2.merge_cells('G4:G6')

            header_fill = PatternFill('solid', fgColor="FFF2CC")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Định dạng ô trong sheet "Hướng dẫn"
            for row in ws2.iter_rows(min_row=12, max_row=12):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="top")
                    cell.font = Font(name='Times New Roman', bold=False, size=11, color="4A86E9")
                    cell.fill = header_fill

            # Định dạng ô trong sheet "Hướng dẫn"
            for row in ws2.iter_rows(min_row=4, max_row=10):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name='Times New Roman', bold=False, size=11, color="000000")
                    cell.border = thin_border

            # Tạo phản hồi trả về tệp Excel
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'
            wb.save(response)
            return response
        except Exception as e:
            return convert_response(str(e), 400)


class UploadFileImport(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        workspace = request.POST.get('workspace')
        excel_file = request.FILES['file']

        r = random.randint(100000, 999999)
        file_name = f"{r}.xlsx"
        file = ContentFile(excel_file.read(), name=file_name)
        url_file = AwsS3.upload_file(file, 'import_customer_file/')

        customer_import = CustomerImport.objects.create(
            file_url=url_file,
            file_name=file_name,
            status=CustomerImport.Status.IN_PROCESS,
            workspace_id=workspace,
            created_by=user,
        )

        # Mở file Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        customer_duplicate = 0
        customer_err = 0
        customer_success = 0
        customer_total = 0
        # Xử lý dữ liệu từ file Excel
        data = []
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            phone = row[2].value
            if not phone:
                continue
            if len(row[2].value) == 10:
                customer_total += 1
                customer = Customer()
                try:
                    customer = Customer.objects.get(phone=phone)
                except Exception as e:
                    customer = Customer.objects.create(
                        phone=phone,
                        prefix_name=row[1].value,
                        # birthday=datetime.strptime(row[3].value, "%d/%m/%Y").strftime("%Y-%m-%d") if row[3].value else None,
                        address=row[4].value,
                        email=row[5].value,
                        gender='MALE' if row[6].value == 1 else 'FEMALE',
                        file_import=customer_import,
                        workspace_id=workspace,
                        created_by=user
                    )
                finally:
                    for row_tag in ws.iter_rows(min_row=5, max_row=5, min_col=8, max_col=ws.max_column + 1):
                        is_duplicate = True
                        for i in list(range(0, ws.max_column - 8)):
                            id_tag = row_tag[i].value
                            try:
                                if row[row_tag[i].col_idx - 1].value == 1.0:
                                    tag = Tag.objects.get(id=id_tag)
                                    cuz = CustomerUserZalo.objects.filter(customer=customer, oa=tag.oa).first()
                                    if not cuz:
                                        CustomerUserZalo.objects.create(
                                            customer=customer,
                                            oa=tag.oa
                                        )
                                        is_duplicate = False
                                    TagCustomer.objects.create(
                                        created_by=user,
                                        customer=customer,
                                        tag=tag,
                                    )
                            except Exception as e:
                                print(str(e))
                        if is_duplicate:
                            customer_duplicate += 1
                        else:
                            customer_success += 1

        customer_import.customer_total = customer_total
        customer_import.customer_success = customer_success
        customer_import.customer_double = customer_duplicate
        customer_import.status = CustomerImport.Status.SUCCESS
        customer_import.save()

        return convert_response('success', 200, data={
            "id": customer_import.id,
            "customer_total": customer_import.customer_total,
            "customer_double": customer_import.customer_double,
            "customer_success": customer_import.customer_success,
        })
